import json
import sys
from openpyxl import load_workbook

if (len(sys.argv) < 2):
  print("Excelファイル名を指定してください。")
  sys.exit()

file_name = sys.argv[1]

workbook = load_workbook(filename=file_name, data_only=True)

##########################################################
sheet = workbook["文言定義"]

WEB_KEY_INDEX = 5 # F列 開発用キー
WEB_VALUE_INDEX = 3 # D列 文言

MOBILE_KEY_INDEX = 6 # G列 開発用キー(Mobile)
MOBILE_VALUE_INDEX = 4 # E列 モバイルアプリ専用文言（Webと異なる場合に記載）

labels_web = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
  if row[WEB_KEY_INDEX] is not None and row[WEB_VALUE_INDEX] is not None:
    key = row[WEB_KEY_INDEX].replace(" ", "").replace("\t", "").replace("\n", "")
    labels_web[key] = row[WEB_VALUE_INDEX]

json_data_web = json.dumps(labels_web, indent=4, ensure_ascii=False, sort_keys=True)

with open("label.ja.json", "w", encoding="utf-8") as f:
  f.write(json_data_web)

print("文言定義を label.ja.json に出力しました。")


labels_mobile = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
  key = row[MOBILE_KEY_INDEX].replace(" ", "").replace("\t", "").replace("\n", "")

  if key is not None:
    if row[MOBILE_VALUE_INDEX] is not None:
      labels_mobile[key] = row[MOBILE_VALUE_INDEX]
    elif row[WEB_VALUE_INDEX] is not None:
      labels_mobile[key] = row[WEB_VALUE_INDEX]
    else:
      continue

json_data_mobile = json.dumps(labels_mobile, indent=4, ensure_ascii=False, sort_keys=True)

with open("mobile-label.ja.json", "w", encoding="utf-8") as f:
  f.write(json_data_mobile)

print("文言定義を mobile-label.ja.json に出力しました。")
##########################################################


##########################################################
sheet_messages = workbook["メッセージ定義"]
MESSAGE_KEY_INDEX = 10 # K列 開発用キー
MESSAGE_VALUE_INDEX = 8 # I列 文言

sheet_validation_messages = workbook["バリデーションメッセージ定義"]
VALIDATION_MESSAGE_KEY_INDEX = 4 # E列 開発用キー
VALIDATION_MESSAGE_VALUE_INDEX = 3 # D列 文言

messages = {}
for row in sheet_messages.iter_rows(min_row=2, values_only=True):
  if row[MESSAGE_KEY_INDEX] is not None and row[MESSAGE_VALUE_INDEX] is not None:
    key = row[MESSAGE_KEY_INDEX].replace(" ", "").replace("\t", "").replace("\n", "")
    messages[key] = row[MESSAGE_VALUE_INDEX]

for row in sheet_validation_messages.iter_rows(min_row=2, values_only=True):
  if row[VALIDATION_MESSAGE_KEY_INDEX] is not None and row[VALIDATION_MESSAGE_VALUE_INDEX] is not None:
    key = row[VALIDATION_MESSAGE_KEY_INDEX].replace(" ", "").replace("\t", "").replace("\n", "")
    messages[key] = row[VALIDATION_MESSAGE_VALUE_INDEX]

json_data_messages = json.dumps(messages, indent=4, ensure_ascii=False, sort_keys=True)
with open("message.ja.json", "w", encoding="utf-8") as f:
  f.write(json_data_messages)

print("メッセージ定義とバリデーションメッセージ定義を message.ja.json に出力しました。")
##########################################################
